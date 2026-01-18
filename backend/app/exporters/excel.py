from openpyxl import load_workbook
from openpyxl.styles import numbers
import pandas as pd

def df_to_excel(
    df: pd.DataFrame,
    path,
    currency_columns: list[str] | None = None,
):
    # 1) escribir DF
    df.to_excel(path, index=False)

    if not currency_columns:
        return

    # 2) abrir workbook para formatear
    wb = load_workbook(path)
    ws = wb.active

    # encabezados
    headers = [cell.value for cell in ws[1]]

    for col_name in currency_columns:
        if col_name not in headers:
            continue

        col_idx = headers.index(col_name) + 1  # 1-based

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)

            if isinstance(cell.value, (int, float)):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(path)
